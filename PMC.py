# -*- coding: cp1252 -*-
import numpy as np
import openpyxl
import time
import normaliza
from scipy.io import savemat

# ESTRUTURA: 4,10,2

# 1passo - definindo o tipo de estrutura e importando 

from pybrain.structure import FeedForwardNetwork,RecurrentNetwork
n = FeedForwardNetwork()


# 2passo - criando e inserindo as camadas na rede feed foward

from pybrain.structure import LinearLayer, SigmoidLayer, BiasUnit, TanhLayer

entrada= LinearLayer(5, name='ENTRY_LAYER')
camada_escondida = TanhLayer(10, name='HIDDEN1_LAYER') #SigmoidLayer(2)
#camada_escondida2 =TanhLayer(10, name='HIDDEN2_LAYER')
saida = LinearLayer(2, name='OUT_LAYER')
bias1= BiasUnit()
bias2 = BiasUnit()
#bias3= BiasUnit()

n.addInputModule(entrada)
n.addModule(camada_escondida)
#n.addModule(camada_escondida2) #cuidado. o modulo é input ,output ou module
n.addModule(bias1)
n.addModule(bias2)
#n.addModule(bias3)
n.addOutputModule(saida)


# 3passo - fazendo as ligações full connection das camadas
from pybrain.structure import FullConnection

h1= FullConnection(entrada,camada_escondida)
#h2 = FullConnection(camada_escondida,camada_escondida2)
c2 = FullConnection(camada_escondida,saida)
bias1_h1= FullConnection(bias1,camada_escondida)
#bias2_h2= FullConnection(bias2,camada_escondida2)
bias3_c2= FullConnection(bias2,saida)


# 4passo - adicionar a rede e finalizar   
n.addConnection(h1)
#n.addConnection(h2)
n.addConnection(c2)
n.addConnection(bias1_h1)
#n.addConnection(bias2_h2)
n.addConnection(bias3_c2)


#5 passo- setando a rede
n.sortModules()

#print n.params


############################### FASE DE TREINAMENTO ############

# 6 passo - arrumando o dataset
from pybrain.datasets import SupervisedDataSet

book=openpyxl.load_workbook('normalizado.xlsx')
sheet=book.active

num_entradas=sheet.max_row-1
num_amostras=sheet.max_column-1
#num_amostras_treino=num_amostras-540


x_amostras=[]   #matrix of amostras
x=None
x_1=None
for i in range(2,num_amostras+3): #coluna
    x_amostras.append(x)
    del x
    x=[]
    for j in range(2,num_entradas):  ##linha
        x_1=sheet.cell(row=j,column=i).value
        if(isinstance(x_1,long)==True):
            x_1=float(x_1)
        x.append(x_1)
        
del x_amostras[0]
del x
del x_1


'''
x_amostras_teste=[]   #matrix of amostras
x=None
x_1=None
for i in range(num_amostras_treino+2,num_amostras+3): #coluna
    x_amostras_teste.append(x)
    del x
    x=[]
    for j in range(2,num_entradas):  ##linha
        x_1=sheet.cell(row=j,column=i).value
        if(isinstance(x_1,long)==True):
            x_1=float(x_1)
        x.append(x_1)
        
del x_amostras_teste[0]      
'''

d_saidas=[]   #matrix of amostras
x=None
x_1=None
for i in range(2,num_amostras+3): #coluna
    d_saidas.append(x)
    del x
    x=[]
    for j in range(num_entradas,num_entradas+2):  ##linha
        x_1=sheet.cell(row=j,column=i).value
        if(isinstance(x_1,long)==True):
            x_1=float(x_1)
        
        x.append(x_1)
        
del d_saidas[0]
del x
del x_1



ds = SupervisedDataSet(5, 2)  #sempre observar as entradas e os target

for i in range(0,num_amostras):    
    ds.addSample((x_amostras[i][0],x_amostras[i][1],x_amostras[i][2],x_amostras[i][3],x_amostras[i][4]),(d_saidas[i][0],d_saidas[i][1]))
#print ds
#print len(ds) 


# 7 passo - treinando
from pybrain.supervised.trainers import BackpropTrainer

trainer = BackpropTrainer(n,ds, learningrate=0.12, momentum=0.05)




for i in xrange(10):
    trainer.train()
    

print(trainer.train())  #MEAN SQUARE ERROR




print 'PARTE DO TESTEEEE'
#X=np.linspace(0,1.5,2000)
#X_max=np.amax(X)
#X_min=np.amin(X)
#X=normaliza.normalizacao(X,'TangH',X_min,X_max)

book=openpyxl.load_workbook('TESTE.xlsx')
sheet=book.active

num_entradas=sheet.max_row-1
num_amostras=sheet.max_column-1

x_amostras_teste=[]   #matrix of amostras
x=None
x_1=None
for i in range(2,num_amostras+3): #coluna
    x_amostras_teste.append(x)
    del x
    x=[]
    for j in range(2,num_entradas):  ##linha
        x_1=sheet.cell(row=j,column=i).value
        if(isinstance(x_1,long)==True):
            x_1=float(x_1)
        x.append(x_1)
        
del x_amostras_teste[0]      



max_i= 0.93815
min_i= -0.94868      #NORMALIZAÇAO SAIDA 
max_j= 0.94733
min_j= -0.94291


max_1=0.678300
min_1=0.085431
max_2=0.90946
min_2=0.049587
max_3=0.67179
min_3=0.097326        #NORMALIZAÇÃO ENTRADA
max_4=0.91227
min_4=0.043315


#v1= normaliza.desnormaliza(aux_i,'TangH',min_1,max_1)




y_resp=[]
for a,b,c,d,e in x_amostras_teste:
    #print(a)
    #print(b)
    ##print(c)
    #print(d)
    #print(e)
    
    va= normaliza.normalizacao(a,'TangH',min_1,max_1)
    vb= normaliza.normalizacao(b,'TangH',min_2,max_2)
    vc= normaliza.normalizacao(c,'TangH',min_3,max_3)
    vd= normaliza.normalizacao(d,'TangH',min_4,max_4)
    [aux_i,aux_j]=n.activate((va,vb,vc,vd,e))
   
    #print(aux_i)
    #print(aux_j)
    aux_i=normaliza.desnormaliza(aux_i,'TangH',min_i,max_i)
    aux_j=normaliza.desnormaliza(aux_j,'TangH',min_j,max_j)
    #print(aux_i)
    #print(aux_j)
    aux=[aux_i,aux_j]
    #print('________________________________________')
    time.sleep(1)
    
    y_resp.append(aux)


y_resp = {"y": y_resp,}     
savemat("RNA.mat", y_resp)
#sein.plotamigo(y_resp,X_desnorm)
#sein.plotamigo(y_resp,X)

#exit()
#n.params






















